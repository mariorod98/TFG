# -*- coding: utf-8 -*-
"""
Funciones para crear un archivo xls
Autor: Mario Rodríguez Chaves
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

import os.path
import numpy as np

import globales as glo
import funciones_auxiliares as aux

colores = ['ef280f', 'e36b2c', 'e7d40a', '6dc36d', 'ea9999', 'ffe599', 'b6d7a8', 'a2c4c9', 'd199ea', 'eac099', '999eea',
           'ea99c7', 'ffffff']
letras = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'


# Función que abre un workbook de excel
# Parámetros de entrada:
#   archivo: path del archivo xls a abrir
#   titulo: nombre de la worksheet
#   sobreescribir: booleano que indica si se quiere sobreescribir el contenido del archivo
# Return: objeto workbook con el archivo cargado
def abre_workbook(archivo, titulo, sobreescribir):
    # si el archivo existe y no se quiere sobreescribir, carga contenido del archivo
    if os.path.isfile(archivo) and not sobreescribir:
        wb = load_workbook(archivo)
    else:  # en otro caso, se crea un workbook nuevo y se elimina la sheet por defecto
        wb = Workbook()
        wb.remove(wb.active)

    # se crea la sheet con el titulo especifico
    wb.create_sheet(titulo)

    # se devuelve el workbook
    return wb


# Función que devuelve la posición dada una fila y una columna
# Parámetros de entrada:
#   fila: número de fila de la celda
#   columna: número de columna de la celda
# Return: string con formato '0A' con la fila y columna correspondientes
def calcula_posicion(fila, columna):
    return letras[columna] + str(fila)


# Función que dibuja la configuración de la solución en una sheet del workbook
# Parámetros de entrada:
#   solucion: nparray con los valores a representar
#   archivo: path del archivo xls a abrir
#   titulo: nombre de la worksheet
#   sobreescribir: booleano que indica si se quiere sobreescribir el contenido del archivo
def exportar_solucion(solucion, archivo, titulo, sobreescribir):
    # se abre el workbook y se escoge la última página
    wb = abre_workbook(archivo, titulo, sobreescribir)
    ws = wb.worksheets[-1]

    # se inicializan las posiciones
    columna = 0
    fila = 1
    i_ant = 0
    i = 0

    # para cada elemento de la solución
    while i < len(solucion):
        # si se trata de un tren distinto, se pasa a la siguiente fila
        if glo.HORARIO_TRENES[i]['TREN'] != glo.HORARIO_TRENES[i_ant]['TREN']:
            fila += 1
            columna = 1

        # se obtiene la posición a partir de la fila y la columna

        # se asigna a esa posicion la hora de llegada y el color del servicio
        ws.cell(fila, columna, glo.HORARIO_TRENES[i]['TIEMPO'])
        ws.cell(fila, columna).fill = PatternFill(fill_type='solid',
                                                  start_color=colores[solucion[i]],
                                                  end_color=colores[solucion[i]])

        # se aumentan los contadores
        columna += 1
        i_ant = i
        i += 1

    # se guarda el workbook
    wb.save(archivo)


def crea_grafico(valores, titulo, y_axis, x_axis):
    # se añade al workbook la gráfica de la evolución del fitness
    values = valores
    chart = LineChart()
    chart.title = titulo
    chart.style = 13
    chart.y_axis.title = y_axis
    chart.x_axis.title = x_axis
    chart.add_data(values)

    return chart


# Función que crea gráficos de progreso por iteración del valor de la función fitness
# la media de tiempos de trabajo y la media de tiempos de descanso
# Parámetros de entrada:
#   titulo_grafico: título de los gráficos
#   archivo: path del archivo xls a abrir
#   titulo: nombre de la worksheet
#   sobreescribir: booleano que indica si se quiere sobreescribir el contenido del archivo
def aniade_graficos(titulo_grafico, archivo, titulo, sobreescribir, es=False):
    # se abre el workbook y se escoge la última página
    wb = abre_workbook(archivo, titulo, sobreescribir)
    ws = wb.worksheets[-1]

    # se obtiene una lista con los números de iteración
    iteraciones = [it for it, fit, resultados, sol in glo.RESULTADOS]
    # se obtiene una lista con el valor fitness de cada iteración
    fitness = [fit for it, fit, resultados, sol in glo.RESULTADOS]

    # se obtiene una lista con los tiempos de trabajo de cada iteración
    t_trabajo = [np.array(resultados['T_TRABAJO'], dtype=int) for it, fit, resultados, sol in glo.RESULTADOS]
    # se obtiene una lista con los tiempos de descanso de cada iteración
    t_descanso = [np.array(resultados['T_DESCANSO'], dtype=int) for it, fit, resultados, sol in glo.RESULTADOS]
    # se obtiene una lista con la cantidad de periodos de trabajo en cada iteración
    n_periodos = [np.array(resultados['N_PERIODOS'], dtype=int) for it, fit, resultados, sol in glo.RESULTADOS]
    # se obtiene una lista con los tiempos de jonada de cada iteración
    t_jornada = [np.array(resultados['T_JORNADA'], dtype=int) for it, fit, resultados, sol in glo.RESULTADOS]

    if es:
        temperatura = [resultados['temperatura'] for it, fit, resultados, sol in glo.RESULTADOS]
        aceptacion = [resultados['aceptacion'] for it, fit, resultados, sol in glo.RESULTADOS]

    # para cada iteración, se calcula la media de tiempos de trabajo
    mean_trabajo = [t.mean() for t in t_trabajo]
    # para cada iteración, se calcula la media de los tiempos de descanso
    mean_descanso = [t.mean() for t in t_descanso]
    # para cada iteración, se calcula la media de los tiempos de descanso
    mean_periodos = [t.mean() for t in n_periodos]
    # para cada iteración, se calcula la media de los tiempos de descanso
    mean_jornada = [t.mean() for t in t_jornada]

    std_trabajo = [t.std() for t in t_trabajo]
    std_descanso = [t.std() for t in t_descanso]

    if es:
        # se almacenan las iteraciones en el workbook
        for i in range(0, len(iteraciones)):
            ws.append([iteraciones[i], fitness[i], mean_trabajo[i], mean_descanso[i], std_trabajo[i], std_descanso[i],
                       mean_periodos[i], mean_jornada[i], temperatura[i], aceptacion[i]])
    else:
        # se almacenan las iteraciones en el workbook
        for i in range(0, len(iteraciones)):
            ws.append([iteraciones[i], fitness[i], mean_trabajo[i], mean_descanso[i], std_trabajo[i], std_descanso[i],
                       mean_periodos[i], mean_jornada[i]])

    valores = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(iteraciones))
    chart = crea_grafico(valores, titulo_grafico, 'Fitness', 'Iteración')
    ws.add_chart(chart, "L1")

    valores = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=len(iteraciones))
    chart = crea_grafico(valores, titulo_grafico, 'Media trabajo', 'Iteración')
    ws.add_chart(chart, "L16")

    valores = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=len(iteraciones))
    chart = crea_grafico(valores, titulo_grafico, 'Media descanso', 'Iteración')
    ws.add_chart(chart, "L31")

    valores = Reference(ws, min_col=5, min_row=1, max_col=5, max_row=len(iteraciones))
    chart = crea_grafico(valores, titulo_grafico, 'Std trabajo', 'Iteración')
    ws.add_chart(chart, "L46")

    valores = Reference(ws, min_col=6, min_row=1, max_col=6, max_row=len(iteraciones))
    chart = crea_grafico(valores, titulo_grafico, 'Std descanso', 'Iteración')
    ws.add_chart(chart, "L61")

    valores = Reference(ws, min_col=7, min_row=1, max_col=7, max_row=len(iteraciones))
    chart = crea_grafico(valores, titulo_grafico, 'Media periodos', 'Iteración')
    ws.add_chart(chart, "L76")

    valores = Reference(ws, min_col=8, min_row=1, max_col=8, max_row=len(iteraciones))
    chart = crea_grafico(valores, titulo_grafico, 'Media jornada', 'Iteración')
    ws.add_chart(chart, "L91")

    valores = Reference(ws, min_col=8, min_row=1, max_col=8, max_row=len(iteraciones))
    chart = crea_grafico(valores, titulo_grafico, 'Media jornada', 'Iteración')
    ws.add_chart(chart, "L91")

    if es:
        valores = Reference(ws, min_col=9, min_row=1, max_col=9, max_row=len(iteraciones))
        chart = crea_grafico(valores, titulo_grafico, 'Temperatura', 'Iteración')
        ws.add_chart(chart, "V1")

        valores = Reference(ws, min_col=10, min_row=1, max_col=10, max_row=len(iteraciones))
        chart = crea_grafico(valores, titulo_grafico, 'Aceptacion', 'Iteración')

        s1 = chart.series[0]
        s1.marker.symbol = "triangle"
        s1.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
        s1.marker.graphicalProperties.line.solidFill = "FF0000"  # Marker outline
        s1.graphicalProperties.line.noFill = True

        ws.add_chart(chart, "V16")

    # se guarda el workbook
    wb.save(archivo)


# Función que transforma un tiempo en minutos en hora y minutos
# Parámetros de entrada:
#   minutos: tiempo en minutos a transformar
# Return: string con formato hh:mm
def min_to_hora(minutos):
    return str((minutos // 60) % 24).rjust(2, '0') + ':' + str(minutos % 60).rjust(2, '0')


# Función que devuelve la columna de excel correspondiente al tiempo dado
# se busca la casilla cuyo tiempo sea más cercano a t_act, sin pasarse
# Parámetros de entrada:
#   t_act: tiempo en minutos que se quiere encasillar
#   tiempos: array con los distintos periodos de tiempo en minutos
# Return: índice del array tiempos con la casilla a la que pertenece t_act
def busca_hora(t_act, tiempos):
    mas_cercano = min(tiempos, key=lambda x: abs(x - t_act))
    if t_act - mas_cercano < 0:
        return tiempos.index(mas_cercano)
    else:
        return tiempos.index(mas_cercano) + 1


# Función que crea el horario de una solución dada
# Parámetros de entrada:
#   solucion: nparray con los valores a representar
#   archivo: path del archivo xls a abrir
#   titulo: nombre de la worksheet
#   sobreescribir: booleano que indica si se quiere sobreescribir el contenido del archivo
def crear_hoja_servicios(solucion, archivo, titulo, sobreescribir):
    # se abre el workbook y se escoge la última página
    wb = abre_workbook(archivo, titulo, sobreescribir)
    ws = wb.worksheets[-1]

    # se crean las listas de tiempos en minutos y en formato hh:mm
    tiempos = []
    horas = []
    # el horario son 24h desde las 4:00 am de un día a las 4:00 am del siguiente
    for t in range(240, 1681, 30):
        tiempos.append(t)
        horas.append(min_to_hora(t))
        horas.append('')

    # se modifica el tamaño de las columnas del archivo excel
    for col in range(1, len(horas) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 6

    # se insertan las horas en la worksheet
    ws.append([''] + horas)

    # se unen las celdas de la banda de las horas, de modo que cada hora ocupa dos casillas
    for i in range(0, len(horas) * 2, 2):
        ws.merge_cells(start_row=1, start_column=i + 2, end_row=1, end_column=i + 3)

    # para cada servicio
    for servicio in range(0, glo.N_SERVICIOS):
        # se fija la fila del servicio
        fila = servicio + 2
        # se indica el servicio en la primera columna
        ws.cell(fila, 1, 'Servicio ' + str(servicio))

        # se calculan los periodos
        periodos = aux.calcula_periodos_servicio(solucion, servicio)

        # para cada periodo
        for p in periodos:
            # se obtienen los tiempos de inicio y fin del periodo
            t_ini = glo.HORARIO_TRENES[p[0]]['TIEMPO']
            t_fin = glo.HORARIO_TRENES[p[1]]['TIEMPO']

            # se obtiene el tren del periodo
            tren = glo.HORARIO_TRENES[p[1]]['TREN']

            # se obtienen las posiciones en el excel del periodo
            # se multiplica por 2 porque las horas ocupan dos celdas
            pos_ini = busca_hora(t_ini, tiempos) * 2
            pos_fin = busca_hora(t_fin, tiempos) * 2

            # si el periodo es de una única franja de tiempo, se almacena en la misma casilla inicio y fin
            if pos_ini == pos_fin:
                ws.cell(fila, pos_ini).value = 'tren ' + str(tren) + ' ' + \
                                               min_to_hora(t_ini) + '-' + min_to_hora(t_fin)
                ws.cell(fila, pos_ini).fill = PatternFill(fill_type='solid',
                                                          start_color=colores[tren],
                                                          end_color=colores[tren])
            else: # si son de distintas franjas de tiempo
                # se pone en la casilla de inicio del periodo el tren y la fecha de inicio
                ws.cell(fila, pos_ini + 1, 'tren ' + str(tren) + ' ' + min_to_hora(t_ini))
                # se pone en la casilla de fin del periodo el tren y la fecha de fin
                ws.cell(fila, pos_fin, min_to_hora(t_fin))

                # se pintan todas las casillas entre la inicial y la final con el color del tren correspondiente
                for i in range(pos_ini + 1, pos_fin + 1):
                    ws.cell(fila, i).fill = PatternFill(fill_type='solid',
                                                        start_color=colores[tren],
                                                        end_color=colores[tren])

    # se guarda el workbook
    wb.save(archivo)
